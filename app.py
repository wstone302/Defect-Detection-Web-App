from flask import Flask, render_template, request, send_file, jsonify
import torch
import os
import cv2
import numpy as np
from pathlib import Path
from datetime import datetime
from PIL import Image
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 英文類別 ➝ 繁體中文
LABEL_MAPPING = {
    "rust": "鏽蝕",
    "peeling_paint": "漆面剝落",
    "rupture": "斷裂",
    "burnt": "燒焦",
    "expansion": "膨脹",
    "missing": "缺件",
    "oil_leakage": "漏油",
    "unknown_object": "異常物"
}

# 載入模型
model = torch.hub.load('yolov5', 'custom', path='best.pt', source='local')
model.conf = 0.25  # 信心閾值

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/detect', methods=['POST'])
def detect():
    file = request.files['file']
    if not file:
        return jsonify({'error': '未上傳任何檔案'}), 400

    filename = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    file.save(filepath)

    results = model(filepath)
    results.render()

    names = model.names
    detections = results.pred[0].cpu().numpy()
    class_counts = {}
    for det in detections:
        cls_id = int(det[5])
        name = names[cls_id]
        class_counts[name] = class_counts.get(name, 0) + 1

    img = Image.fromarray(results.ims[0])
    result_path = os.path.join(UPLOAD_FOLDER, 'result_' + filename)
    os.makedirs(os.path.dirname(result_path), exist_ok=True)
    img.save(result_path)

    return jsonify({
        'image_path': result_path,
        'class_counts': class_counts
    })

@app.route('/video_detect', methods=['POST'])
def video_detect():
    file = request.files['file']
    if not file or not file.filename.endswith(('.mp4', '.avi', '.mov', '.MOV')):
        return jsonify({'error': '請上傳影片檔（.mp4/.avi/.mov）'}), 400

    filename = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + file.filename
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    cap = cv2.VideoCapture(filepath)
    fps = cap.get(cv2.CAP_PROP_FPS)

    all_classes = {}
    image_info_list = []

    frame_idx = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        timestamp = frame_idx / fps
        results = model(frame)
        detections = results.pred[0].cpu().numpy()

        if len(detections) > 0:
            results.render()
            rendered = cv2.cvtColor(results.ims[0], cv2.COLOR_BGR2RGB)
            img_pil = Image.fromarray(rendered)
            img_pil = img_pil.resize((350, 260))

            base_name = os.path.splitext(filename)[0]
            img_name = f"{base_name}_frame_{frame_idx}.jpg"
            img_path = os.path.join(UPLOAD_FOLDER, img_name)
            img_pil.save(img_path)

            labels = []
            for det in detections:
                cls_id = int(det[5])
                name = model.names[cls_id]
                all_classes[name] = all_classes.get(name, 0) + 1
                labels.append(name)

            image_info_list.append({
                "image_path": img_path,
                "timestamp": f"{int(timestamp // 60):02d}:{int(timestamp % 60):02d}",
                "filename": os.path.splitext(file.filename)[0],
                "label": "、".join(sorted(set(labels)))
            })

        frame_idx += 1

    cap.release()

    excel_path = os.path.join(UPLOAD_FOLDER, f'report_{filename}.xlsx')
    if image_info_list:
        generate_excel_with_images(image_info_list, excel_path)

    return jsonify({
        'class_counts': all_classes,
        'excel_path': excel_path
    })

def generate_excel_with_images(image_info_list, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "結果"

    # 設定欄寬
    for col in range(1, 50):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # 設定框線樣式
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    col_offset = 1  # 從第1欄開始
    row = 1         # 從第1列開始

    for idx, info in enumerate(image_info_list):
        img = XLImage(info["image_path"])
        img.width = 450
        img.height = 300

        # 設定圖片起始欄位
        col = col_offset if idx % 2 == 0 else col_offset + 8
        cell_position = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell_position)

        # 設定圖片區的列高，避免圖片超出
        for r in range(row, row + 10):  # ↓↓↓ 修改這邊 ↓↓↓
            ws.row_dimensions[r].height = 23

        # 說明列
        text_row = row + 10  # ↓↓↓ 修改這邊 ↓↓↓
        ws.merge_cells(start_row=text_row, start_column=col, end_row=text_row, end_column=col + 4)
        desc_cell = ws.cell(row=text_row, column=col)

        # 中文轉換
        zh_labels = "、".join([LABEL_MAPPING.get(label, label) for label in info["label"].split("、")])
        desc_cell.value = f'{info["filename"]} {info["timestamp"]} {zh_labels}'
        desc_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 加上外框線（圖片區 + 說明列）
        for r in range(row, text_row + 1):
            for c in range(col, col + 5):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    _ = cell.value

                top = thin_border.top if r == row or r == text_row else None
                bottom = thin_border.bottom if r == text_row else None
                left = thin_border.left if c == col else None
                right = thin_border.right if c == col + 4 else None

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

        # 換下一列（雙欄模式）
        if idx % 2 == 1:
            row += 13  # ↓↓↓ 修改這邊 ↓↓↓（10圖 + 1說明 + 2空白）


    wb.save(output_path)

@app.route('/detect_frame', methods=['POST'])
def detect_frame():
    file = request.files['file']
    if not file:
        return 'No frame received', 400

    nparr = np.frombuffer(file.read(), np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    results = model(frame)
    results.render()

    # 儲存渲染圖，提供給前端 canvas 顯示
    filename = datetime.now().strftime("%Y%m%d_%H%M%S_frame.jpg")
    result_path = os.path.join(UPLOAD_FOLDER, filename)
    img = Image.fromarray(cv2.cvtColor(results.ims[0], cv2.COLOR_BGR2RGB))
    img.save(result_path)

    # 統計類別數量
    detections = results.pred[0].cpu().numpy()
    class_counts = {}
    for det in detections:
        cls_id = int(det[5])
        name = model.names[cls_id]
        class_counts[name] = class_counts.get(name, 0) + 1

    return jsonify({
        "image_path": result_path,
        "class_counts": class_counts
    })

@app.route('/get_image')
def get_image():
    path = request.args.get("path")
    return send_file(path, mimetype='image/jpeg')

@app.route('/get_file')
def get_file():
    path = request.args.get("path")
    return send_file(path, as_attachment=True)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    log = data.get("log", [])

    if not log:
        return jsonify({"error": "No data"}), 400

    records = []
    for item in log:
        for cls, count in item["classes"].items():
            records.append({
                "影片名稱": item["filename"],
                "時間戳": item["timestamp"],
                "類別": cls,
                "數量": count
            })

    df = pd.DataFrame(records)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(UPLOAD_FOLDER, f"detect_log_{timestamp}.xlsx")
    df.to_excel(excel_path, index=False)

    return jsonify({"excel_path": excel_path})

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5002)
